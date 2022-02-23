from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
ws = wb.active # 현재 활성화된 sheet 가져옴

# 제목 행 입력
ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트",])

# 변수에 넣어서 입력(리스트로 묶고 각각을 튜플 처리)
scores=[
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)]

# 반복문으로 값 입력
for s in scores:
    ws.append(s)

# # 퀴즈2 점수를 10으로 수정
# for i in range(2,12):
#    ws.cell(column=4,row=i).value=10 

# 퀴즈2 점수를 10으로 수정(답)
for idx, cell in enumerate(ws["D"]):
    if idx == 0: # 제목인 경우 skip
        continue
    cell.value = 10

# H열에 총점(SUM 이용), I열에 성적 정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    # ws["H2"] = "=SUM(B2:G2)"
    # ws["H3"] = "=SUM(B3:G3)"
    # ws["H4"] = "=SUM(B4:G4)"
    # ws["H5"] = "=SUM(B5:G5)"
    # ws["H6"] = "=SUM(B6:G6)"
    # ws["H7"] = "=SUM(B7:G7)"
    # ws["H8"] = "=SUM(B8:G8)"
    # ws["H9"] = "=SUM(B9:G9)"
    # ws["H10"] = "=SUM(B10:G10)"
    # ws["H11"] = "=SUM(B11:G11)" 

#   - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >=80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

# 출석이 5 미만인 학생은 총점 상관없이 F
    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade # I열에 성적 정보 추가

# 엑셀 저장
wb.save("score.xlsx")
